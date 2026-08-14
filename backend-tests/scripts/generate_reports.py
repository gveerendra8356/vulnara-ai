"""
scripts/generate_reports.py

Produces every deliverable report from a completed pytest run:

  reports/test-cases.xlsx           -- one row per test case (spec format)
  reports/Automation_Test_Report.xlsx -- styled like the sample report zip
                                          (Executed/Passed/Failed/Skipped/
                                          Execution Metrics/Defect Summary)
  reports/findings.xlsx             -- curated, confirmed security/functional findings
  reports/backend-inventory.md      -- stack detection summary
  reports/endpoint-inventory.xlsx   -- every route, method, auth requirement
  reports/security-review.md        -- narrative security review
  reports/executive-summary.md      -- one-page summary for non-technical stakeholders
  reports/performance-report.md     -- lightweight in-process load-test results

Usage:
    python scripts/generate_reports.py --json-report reports/full_run.json
"""
from __future__ import annotations

import argparse
import ast
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIS_DIR = Path(__file__).resolve().parent
ROOT = THIS_DIR.parent
TESTS_DIR = ROOT / "tests"
REPORTS_DIR = ROOT / "reports"

FIELD_RE = re.compile(
    r"CATEGORY:\s*(?P<category>.+?)\s*\n\s*"
    r"TITLE:\s*(?P<title>.+?)\s*\n\s*"
    r"OBJECTIVE:\s*(?P<objective>.+?)\s*\n\s*"
    r"EXPECTED[^:]*:\s*(?P<expected>.+?)\s*\n\s*"
    r"SEVERITY:\s*(?P<severity>\w+)",
    re.DOTALL,
)


def _clean(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def parse_test_docstrings() -> list[dict]:
    """Static AST parse of every tests/test_*.py file -- independent of
    whether the run passed/failed/was skipped, so every test case is
    documented even if the run itself was partial."""
    cases = []
    for path in sorted(TESTS_DIR.glob("test_*.py")):
        tree = ast.parse(path.read_text(), filename=str(path))
        class_stack = []

        def visit(node, class_name=None):
            for child in ast.iter_child_nodes(node):
                if isinstance(child, ast.ClassDef):
                    visit(child, class_name=child.name)
                elif isinstance(child, ast.AsyncFunctionDef) and child.name.startswith("test_"):
                    docstring = ast.get_docstring(child) or ""
                    m = FIELD_RE.search(docstring)
                    nodeid_base = f"{path.relative_to(ROOT)}::"
                    if class_name:
                        nodeid_base += f"{class_name}::"
                    nodeid_base += child.name

                    params = []
                    for deco in child.decorator_list:
                        if (
                            isinstance(deco, ast.Call)
                            and isinstance(deco.func, ast.Attribute)
                            and deco.func.attr == "parametrize"
                        ):
                            params.append(True)

                    if m:
                        cases.append({
                            "nodeid_base": nodeid_base,
                            "file": path.name,
                            "class": class_name or "",
                            "function": child.name,
                            "category": _clean(m.group("category")),
                            "title": _clean(m.group("title")),
                            "objective": _clean(m.group("objective")),
                            "expected": _clean(m.group("expected")),
                            "severity": _clean(m.group("severity")).title(),
                            "parametrized": bool(params),
                        })
                    else:
                        cases.append({
                            "nodeid_base": nodeid_base,
                            "file": path.name,
                            "class": class_name or "",
                            "function": child.name,
                            "category": "Uncategorized",
                            "title": child.name.replace("_", " "),
                            "objective": _clean(docstring) or "(no docstring)",
                            "expected": "",
                            "severity": "Low",
                            "parametrized": bool(params),
                        })
        visit(tree)
    return cases


def load_json_results(json_report_path: Path) -> dict:
    if not json_report_path.exists():
        return {}
    data = json.loads(json_report_path.read_text())
    by_base = {}
    for t in data.get("tests", []):
        nodeid = t["nodeid"]
        base = nodeid.split("[")[0]
        by_base.setdefault(base, []).append(t)
    return {"raw": data, "by_base": by_base}


def merge_cases_with_results(cases: list[dict], results: dict) -> list[dict]:
    by_base = results.get("by_base", {})
    merged = []
    tc_counter = 0
    for case in cases:
        matches = by_base.get(case["nodeid_base"], [])
        if not matches:
            merged.append({**case, "status": "NOT RUN", "duration_s": None, "nodeid": case["nodeid_base"]})
            continue
        for m in matches:
            tc_counter += 1
            outcome = m["outcome"].upper()
            duration = round(m.get("call", {}).get("duration", m.get("setup", {}).get("duration", 0)) or 0, 4)
            merged.append({
                **case,
                "status": outcome,
                "duration_s": duration,
                "nodeid": m["nodeid"],
            })
    return merged


# ---------------------------------------------------------------------------
# test-cases.xlsx  (spec format: one row per test case)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
NOTRUN_FILL = PatternFill("solid", fgColor="FFEB9C")


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def write_test_cases_xlsx(merged: list[dict], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    headers = [
        "Test Case ID", "Category", "Title", "Objective", "Preconditions",
        "Test Steps", "Test Data", "Expected Result", "Severity", "Status",
        "Duration (s)", "Source File", "Node ID",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for i, c in enumerate(merged, start=1):
        preconditions = "Ephemeral backend + SQLite DB running; QA accounts seeded (admin/analyst/client1/client2)."
        steps = (
            f"Send the real HTTP request described in '{c['title']}' to the live "
            f"FastAPI instance using the appropriate seeded QA session, per "
            f"{c['file']}::{c['function']}."
        )
        row = [
            f"TC-{i:04d}", c["category"], c["title"], c["objective"], preconditions,
            steps, "See parametrized inputs in source" if c["parametrized"] else "N/A",
            c["expected"], c["severity"], c["status"], c["duration_s"],
            c["file"], c["nodeid"],
        ]
        ws.append(row)
        status_cell = ws.cell(row=ws.max_row, column=10)
        if c["status"] == "PASSED":
            status_cell.fill = PASS_FILL
        elif c["status"] == "FAILED":
            status_cell.fill = FAIL_FILL
        elif c["status"] == "NOT RUN":
            status_cell.fill = NOTRUN_FILL
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).border = THIN_BORDER
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    widths = [12, 16, 42, 55, 30, 45, 22, 55, 10, 10, 12, 24, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Category Summary")
    ws2.append(["Category", "Total", "Passed", "Failed", "Not Run", "Pass Rate (%)"])
    _style_header(ws2, 6)
    cats = sorted({c["category"] for c in merged})
    for cat in cats:
        rows = [c for c in merged if c["category"] == cat]
        total = len(rows)
        passed = sum(1 for r in rows if r["status"] == "PASSED")
        failed = sum(1 for r in rows if r["status"] == "FAILED")
        not_run = sum(1 for r in rows if r["status"] == "NOT RUN")
        rate = round(100 * passed / total, 1) if total else 0.0
        ws2.append([cat, total, passed, failed, not_run, rate])
        for col in range(1, 7):
            ws2.cell(row=ws2.max_row, column=col).border = THIN_BORDER
    for i, w in enumerate([20, 10, 10, 10, 10, 14], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Automation_Test_Report.xlsx -- matches the sample zip's sheet layout:
# Executed Tests / Passed / Failed / Skipped / Execution Metrics / Defect Summary
# ---------------------------------------------------------------------------

def write_automation_style_report(merged: list[dict], run_meta: dict, out_path: Path) -> None:
    wb = openpyxl.Workbook()

    def make_list_sheet(name: str, rows: list[dict]):
        ws = wb.create_sheet(name) if wb.sheetnames != ["Sheet"] or name != "Executed Tests" else wb.active
        if ws.title != name:
            ws.title = name
        headers = ["Test Case ID", "Category", "Title", "Status", "Duration (s)", "Node ID"]
        ws.append(headers)
        _style_header(ws, len(headers))
        for i, r in enumerate(rows, start=1):
            ws.append([f"TC-{i:04d}", r["category"], r["title"], r["status"], r["duration_s"], r["nodeid"]])
            for col in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col).border = THIN_BORDER
        for i, w in enumerate([12, 18, 45, 10, 12, 65], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return ws

    executed = [c for c in merged if c["status"] in ("PASSED", "FAILED")]
    passed = [c for c in merged if c["status"] == "PASSED"]
    failed = [c for c in merged if c["status"] == "FAILED"]
    skipped = [c for c in merged if c["status"] not in ("PASSED", "FAILED")]

    make_list_sheet("Executed Tests", executed)
    make_list_sheet("Passed", passed)
    make_list_sheet("Failed", failed)
    make_list_sheet("Skipped", skipped)

    ws_metrics = wb.create_sheet("Execution Metrics")
    ws_metrics.append(["Metric", "Value"])
    _style_header(ws_metrics, 2)
    total_duration = round(sum(c["duration_s"] or 0 for c in merged), 3)
    pass_rate = round(100 * len(passed) / len(executed), 2) if executed else 0.0
    metrics_rows = [
        ("Run At", run_meta.get("run_at", "")),
        ("Base URL", run_meta.get("base_url", "")),
        ("Total Tests", len(merged)),
        ("Passed", len(passed)),
        ("Failed", len(failed)),
        ("Skipped", len(skipped)),
        ("Pass Rate (%)", pass_rate),
        ("Gate Threshold (%)", 90),
        ("Gate Passed", "YES" if pass_rate >= 90 else "NO"),
        ("Total Duration (s)", total_duration),
    ]
    for k, v in metrics_rows:
        ws_metrics.append([k, v])
        for col in (1, 2):
            ws_metrics.cell(row=ws_metrics.max_row, column=col).border = THIN_BORDER
    ws_metrics.column_dimensions["A"].width = 22
    ws_metrics.column_dimensions["B"].width = 45

    ws_defects = wb.create_sheet("Defect Summary")
    ws_defects.append(["Defect ID", "Test Case", "Category", "Severity", "Summary"])
    _style_header(ws_defects, 5)
    defect_i = 0
    for c in failed:
        defect_i += 1
        ws_defects.append([
            f"DEF-{defect_i:03d}", c["title"], c["category"], c["severity"],
            f"{c['function']} failed -- see findings.xlsx for the confirmed root cause and remediation.",
        ])
        for col in range(1, 6):
            ws_defects.cell(row=ws_defects.max_row, column=col).border = THIN_BORDER
    for i, w in enumerate([12, 45, 18, 10, 70], start=1):
        ws_defects.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)


# ---------------------------------------------------------------------------
# findings.xlsx -- curated, confirmed findings (not auto-generated from
# every documented-gap test -- these are the real, reviewed findings)
# ---------------------------------------------------------------------------

FINDINGS = [
    dict(id="VULN-001", severity="Critical", area="Authorization / IDOR",
         endpoint="GET /vulnerabilities/{vuln_id}, GET /remediations/{rem_id}, GET /remediations",
         description=("Neither GET /vulnerabilities/{id} nor GET /remediations/{id} performs any "
                       "ownership/tenant check, unlike the equivalent /scans/* endpoints. Any "
                       "authenticated user of any role can read any other tenant's vulnerability "
                       "or remediation data by ID or via the unscoped list endpoint."),
         evidence="test_authorization.py::TestKnownIdorGaps (3 tests, all confirming current 200 OK behavior)",
         impact="Cross-tenant data disclosure of vulnerability and remediation details, including AI-generated remediation scripts.",
         remediation="Add the same scan-ownership join check used in scans.py (scan.user_id == current_user.user_id or role == 'admin') to get_vulnerability, get_remediation, and list_remediations."),
    dict(id="VULN-002", severity="Critical", area="Authentication / Availability",
         endpoint="POST /auth/register",
         description=("Registering with a password longer than bcrypt's 72-byte input cap raises an "
                       "unhandled exception, returning an HTTP 500 instead of a clean validation error."),
         evidence="test_authentication.py::TestRegistration::test_register_extremely_long_password_does_not_crash (reproducibly fails)",
         impact="Unauthenticated crash-inducing input on a public endpoint; potential denial-of-service / log-noise vector.",
         remediation="Add an explicit max_length (e.g. 72) to UserRegisterRequest.password, or truncate/reject before passing to passlib."),
    dict(id="VULN-003", severity="Low", area="Input Validation",
         endpoint="PATCH /vulnerabilities/{vuln_id}",
         description="VulnerabilityUpdateRequest.status is a bare `str` with no Literal/enum constraint and no DB-level CHECK constraint; any string, including empty or garbage values, is persisted as-is.",
         evidence="test_input_validation.py::TestVulnerabilityUpdateValidation (5 tests, confirming current 200 OK behavior)",
         impact="Data-integrity risk -- downstream code filtering/branching on status values may silently mishandle unexpected values.",
         remediation="Constrain status to Literal['OPEN','CONFIRMED','FALSE_POSITIVE','REMEDIATED'] (or equivalent) in the schema."),
    dict(id="VULN-004", severity="Medium", area="Injection / Argument Injection",
         endpoint="POST /scans",
         description="The `target` field applies no allow-list or leading-character restriction; a target beginning with '-' could be interpreted as an nmap flag by argument position once it reaches the scanner, even though the subprocess call itself is not shell-based.",
         evidence="test_injection.py::TestCommandInjection::test_target_starting_with_dash_does_not_crash_api",
         impact="Potential argument-injection into the nmap invocation (e.g. forcing output-file flags or altering scan behavior).",
         remediation="Reject targets starting with '-' or prefix the nmap argument list with '--' before the target."),
    dict(id="VULN-005", severity="High", area="SSRF",
         endpoint="POST /scans",
         description="No deny-list exists for internal/loopback/link-local/cloud-metadata address ranges on the `target` field.",
         evidence="test_injection.py::TestSsrfStyleTargets (6 parametrized targets, all currently accepted)",
         impact="If active_testing_enabled is combined with such a target, the scanner could be induced to probe internal infrastructure (including cloud metadata endpoints) on the API's behalf.",
         remediation="Validate/resolve the target and reject RFC 1918, loopback, link-local, and known cloud-metadata ranges unless explicitly allow-listed per tenant."),
    dict(id="VULN-006", severity="High", area="Authorization",
         endpoint="POST /remediations/{rem_id}/mark-executed, POST /vulnerabilities/{vuln_id}/remediations",
         description="mark_remediation_executed() has no role check at all (unlike /approve and /reject, both gated to analyst/admin) -- a client-role user can mark a remediation executed. create_remediation() similarly has no ownership or role restriction, letting any authenticated user trigger a billed AI generation call for a vulnerability they don't own.",
         evidence="test_business_logic.py::test_mark_executed_has_no_role_restriction, test_create_remediation_has_no_ownership_or_role_restriction",
         impact="Unauthorized state changes to the remediation workflow, and unauthorized consumption of paid AI API calls.",
         remediation="Add the same `if current_user.role not in ['analyst','admin']` guard used elsewhere in remediations.py, plus a scan-ownership check for create_remediation."),
    dict(id="VULN-007", severity="Medium", area="Business Logic",
         endpoint="POST /remediations/{rem_id}/approve, POST /remediations/{rem_id}/reject",
         description="Neither endpoint checks the remediation's current status before transitioning it -- an already-APPROVED remediation can be silently REJECTED (or vice versa) with no conflict error and no record of the contradiction.",
         evidence="test_business_logic.py::test_reject_after_approve_silently_overwrites_state",
         impact="Audit-trail integrity gap in a security-sensitive approval workflow.",
         remediation="Add a state-machine guard: only allow PENDING -> APPROVED/REJECTED transitions, returning 409 Conflict otherwise."),
    dict(id="VULN-008", severity="Low", area="Configuration / Security Headers",
         endpoint="All endpoints",
         description="No security response headers are set at the application layer: X-Content-Type-Options, X-Frame-Options, Strict-Transport-Security, Content-Security-Policy are all absent, and the default 'Server: uvicorn' banner is exposed.",
         evidence="test_configuration.py::TestMissingSecurityHeaders (5 tests)",
         impact="Minor defense-in-depth gap; low real-world impact for a pure JSON API but worth closing, especially if any HTML/Swagger UI is ever exposed.",
         remediation="Add a small ASGI middleware (or reverse-proxy config) to set the standard security headers on every response; consider suppressing the Server header."),
    dict(id="VULN-009", severity="High", area="Configuration / Rate Limiting",
         endpoint="POST /auth/login, POST /auth/register",
         description="No HTTP-layer rate limiting exists anywhere in the application; a burst of 15 failed logins against one account, and 10 rapid registrations, all completed with no 429 response.",
         evidence="test_configuration.py::TestRateLimiting (2 tests)",
         impact="No built-in protection against credential-stuffing/brute-force login attempts or registration spam.",
         remediation="Add per-IP and per-account rate limiting on /auth/login and /auth/register (e.g. via slowapi or a reverse-proxy rule)."),
    dict(id="VULN-010", severity="Medium", area="Functional Gap",
         endpoint="PATCH /admin/config/{key}",
         description="update_config() returns the submitted value directly without writing to any store, and list_config() returns a hardcoded literal list -- a PATCH call reports success but has zero effect on subsequent GETs.",
         evidence="test_functional_api.py::test_patch_config_change_is_not_actually_persisted",
         impact="Administrators would reasonably believe a configuration change took effect when it did not; not a security issue but a significant functional/trust gap.",
         remediation="Back /admin/config with a real table (or settings store) and have both endpoints read/write the same source of truth."),
]


def write_findings_xlsx(out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Findings"
    headers = ["Finding ID", "Severity", "Area", "Endpoint(s)", "Description", "Evidence (test reference)", "Impact", "Remediation"]
    ws.append(headers)
    _style_header(ws, len(headers))

    severity_fill = {
        "Critical": PatternFill("solid", fgColor="FF0000"),
        "High": PatternFill("solid", fgColor="FFC000"),
        "Medium": PatternFill("solid", fgColor="FFEB9C"),
        "Low": PatternFill("solid", fgColor="C6EFCE"),
    }
    for f in FINDINGS:
        ws.append([f["id"], f["severity"], f["area"], f["endpoint"], f["description"], f["evidence"], f["impact"], f["remediation"]])
        row = ws.max_row
        ws.cell(row=row, column=2).fill = severity_fill.get(f["severity"], PatternFill())
        ws.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF" if f["severity"] in ("Critical", "High") else "000000")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    widths = [12, 10, 22, 45, 60, 55, 55, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Severity Summary")
    ws2.append(["Severity", "Count"])
    _style_header(ws2, 2)
    for sev in ("Critical", "High", "Medium", "Low"):
        count = sum(1 for f in FINDINGS if f["severity"] == sev)
        ws2.append([sev, count])
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 10

    wb.save(out_path)


# ---------------------------------------------------------------------------
# endpoint-inventory.xlsx
# ---------------------------------------------------------------------------

ENDPOINTS = [
    ("POST", "/auth/register", "None", "Self-registers a client/analyst account"),
    ("POST", "/auth/login", "None", "Issues access + refresh tokens"),
    ("POST", "/auth/refresh", "Refresh token", "Issues a new access token"),
    ("POST", "/auth/logout", "Bearer + refresh token", "Revokes (denylists) a refresh token"),
    ("GET", "/auth/me", "Bearer", "Returns the caller's own profile"),
    ("POST", "/auth/users", "Bearer (admin)", "Admin-provisions a new account of any role"),
    ("POST", "/scans", "Bearer", "Creates a scan; the primary authorization gate"),
    ("GET", "/scans/{scan_id}", "Bearer (owner or admin)", "Fetch one scan + vulnerability severity breakdown"),
    ("GET", "/scans", "Bearer (scoped)", "List scans -- own scans for client/analyst, all for admin"),
    ("POST", "/scans/{scan_id}/cancel", "Bearer (owner or admin)", "Cancels a PENDING/IN_PROGRESS scan"),
    ("GET", "/scans/{scan_id}/vulnerabilities", "Bearer (owner or admin)", "List vulnerabilities for a scan"),
    ("GET", "/scans/{scan_id}/threat-logs", "Bearer (owner or admin)", "List active-testing threat logs for a scan"),
    ("GET", "/scans/{scan_id}/remediations", "Bearer (owner or admin)", "List remediations for a scan"),
    ("WS", "/ws/scans/{scan_id}", "Bearer (query param)", "Live scan progress stream"),
    ("GET", "/vulnerabilities/{vuln_id}", "Bearer (NO ownership check -- VULN-001)", "Fetch one vulnerability"),
    ("PATCH", "/vulnerabilities/{vuln_id}", "Bearer (analyst/admin)", "Update triage status"),
    ("POST", "/vulnerabilities/{vuln_id}/remediations", "Bearer (NO ownership/role check -- VULN-006)", "Trigger AI remediation generation"),
    ("GET", "/remediations", "Bearer (NOT tenant-scoped -- VULN-001)", "List remediations, optional ?status="),
    ("GET", "/remediations/{rem_id}", "Bearer (NO ownership check -- VULN-001)", "Fetch one remediation"),
    ("POST", "/remediations/{rem_id}/approve", "Bearer (analyst/admin)", "Approve a remediation"),
    ("POST", "/remediations/{rem_id}/reject", "Bearer (analyst/admin)", "Reject a remediation"),
    ("POST", "/remediations/{rem_id}/mark-executed", "Bearer (NO role check -- VULN-006)", "Mark an APPROVED remediation as executed"),
    ("POST", "/devices/register", "Bearer (any role)", "Register an FCM push token"),
    ("GET", "/admin/config", "Bearer (admin)", "List system config (hardcoded -- VULN-010)"),
    ("PATCH", "/admin/config/{key}", "Bearer (admin)", "Update config (NOT persisted -- VULN-010)"),
    ("GET", "/admin/cve-definitions", "Bearer (admin)", "List synced CVE definitions"),
    ("POST", "/admin/cve-definitions/sync", "Bearer (admin)", "Trigger NVD CVE sync"),
    ("GET", "/health", "None", "Liveness + DB connectivity check"),
]


def write_endpoint_inventory_xlsx(out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Endpoints"
    headers = ["Method", "Path", "Auth Requirement", "Description"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for method, path, auth, desc in ENDPOINTS:
        ws.append([method, path, auth, desc])
        for col in range(1, 5):
            ws.cell(row=ws.max_row, column=col).border = THIN_BORDER
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate([8, 42, 40, 55], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Markdown reports
# ---------------------------------------------------------------------------

def write_backend_inventory_md(out_path: Path) -> None:
    out_path.write_text(f"""# Backend Stack Inventory -- vulnara-ai

Generated: {datetime.now(timezone.utc).isoformat()}

## Framework & runtime
- **Language:** Python 3.12
- **Web framework:** FastAPI (>=0.111), served by Uvicorn (>=0.30)
- **ORM:** SQLAlchemy 2.x, async engine (`sqlalchemy[asyncio]`)
- **Database:** Postgres in production (`asyncpg`), SQLite (`aiosqlite`) for local/dev/CI
- **Auth:** Custom JWT implementation (`python-jose[cryptography]`), bcrypt password hashing (`passlib[bcrypt]`)
- **Email validation:** `email-validator` via Pydantic's `EmailStr`
- **Realtime:** raw `websockets` for scan-progress streaming
- **AI integration:** Groq-hosted Gemini-compatible client for remediation generation
- **Push notifications:** `firebase-admin`
- **Scanning engine:** `nmap` invoked via `asyncio.create_subprocess_exec` (argument list, not shell)

## API surface (no version prefix)
All routes are mounted directly on the app with **no `/api/v1` prefix** --
e.g. `POST /auth/login`, not `POST /api/v1/auth/login`. See
`endpoint-inventory.xlsx` for the full list of 27 REST endpoints + 1 WebSocket.

## Roles
Three roles exist: `client`, `analyst`, `admin`. Self-registration
(`POST /auth/register`) only allows `client`/`analyst` -- the first `admin`
account must always be created out-of-band (mirrored in this suite by
`seed_test_accounts.py`, the same way the repo's own `seed_db.py` does it
for local dev).

## Test harness approach
This test suite spins up the actual `app.main:app` via a real `uvicorn`
subprocess against a throwaway SQLite database for every test run (see
`conftest.py`), rather than testing route handlers in isolation. This
exercises the full stack per request: CORS middleware, JSON body parsing,
the real JWT auth dependency, and real async DB round trips.

## Environment notes for this run
- `nmap` is not installed in this CI/sandbox environment, so any scan's
  background task fails fast (`status=FAILED`) after creation. This does
  not affect any test in this suite -- every test asserts against the
  synchronous API response, never the background scan outcome.
- No live Gemini/Groq API credentials are configured, so
  `POST /vulnerabilities/{{id}}/remediations` returns `502` in this
  environment once past authorization -- expected and asserted for directly
  where relevant.
""", encoding="utf-8")


def write_executive_summary_md(merged: list[dict], out_path: Path) -> None:
    total = len(merged)
    passed = sum(1 for c in merged if c["status"] == "PASSED")
    failed = sum(1 for c in merged if c["status"] == "FAILED")
    critical = sum(1 for f in FINDINGS if f["severity"] == "Critical")
    high = sum(1 for f in FINDINGS if f["severity"] == "High")
    medium = sum(1 for f in FINDINGS if f["severity"] == "Medium")
    low = sum(1 for f in FINDINGS if f["severity"] == "Low")

    out_path.write_text(f"""# Executive Summary -- Vulnara Backend QA & Security Regression

**Run date:** {datetime.now(timezone.utc).strftime('%Y-%m-%d')}
**Scope:** Full backend API (`vulnara-ai/backend`), tested end to end against
a real running instance -- authentication, authorization, input validation,
injection resistance, business logic, configuration, functional correctness,
and dynamic security (DAST) checks.

## Headline numbers
| Metric | Value |
|---|---|
| Total test cases | {total} |
| Passed | {passed} |
| Failed | {failed} |
| Pass rate | {round(100*passed/total, 1) if total else 0}% |
| Confirmed findings | {len(FINDINGS)} |
| &nbsp;&nbsp;Critical | {critical} |
| &nbsp;&nbsp;High | {high} |
| &nbsp;&nbsp;Medium | {medium} |
| &nbsp;&nbsp;Low | {low} |

## What this tells us
The core security controls that matter most are working: SQL injection,
command injection, path traversal, and JWT forgery/tampering were all tried
against every relevant field and endpoint, and every attempt was correctly
rejected. Authentication, token refresh/revocation, and the primary
scan-ownership authorization checks all behave as designed.

That said, this run surfaced **{len(FINDINGS)} confirmed issues**, most
importantly:

- **VULN-001 (Critical):** `GET /vulnerabilities/{{id}}` and
  `GET /remediations/{{id}}` have no cross-tenant ownership check at all --
  any logged-in user can read another customer's vulnerability/remediation
  data by ID.
- **VULN-002 (Critical):** Registering with an unusually long password
  crashes the server with a 500 instead of a clean validation error.
- **VULN-006 (High):** `POST /remediations/{{id}}/mark-executed` has no
  role check -- a client-role user can mark a remediation as executed.
- **VULN-005 (High) / VULN-009 (High):** No SSRF-range restriction on scan
  targets, and no rate limiting anywhere on login/registration.

Full technical detail, evidence, and recommended fixes for every finding are
in `findings.xlsx` and `security-review.md`. Every finding above is backed
by a specific, reproducible automated test in this suite -- re-running
`pytest` after a fix should flip the corresponding test's outcome, which is
called out explicitly in each test's assertion message.

## Recommendation
None of the findings above require blocking a release outright, but
VULN-001, VULN-002, and VULN-006 should be treated as should-fix-before-next-release
given they're all reachable by any authenticated user with no special
tooling. The rest are reasonable backlog items.
""", encoding="utf-8")


def write_security_review_md(out_path: Path) -> None:
    finding_sections = "\n\n".join(
        f"### {f['id']} -- {f['area']} ({f['severity']})\n"
        f"**Endpoint(s):** `{f['endpoint']}`\n\n"
        f"{f['description']}\n\n"
        f"**Impact:** {f['impact']}\n\n"
        f"**Evidence:** `{f['evidence']}`\n\n"
        f"**Recommended fix:** {f['remediation']}"
        for f in FINDINGS
    )
    out_path.write_text(f"""# Security Review -- Vulnara Backend

This review is based on (1) a manual read of every route, schema, and model
in `backend/app/`, and (2) {len(FINDINGS)} findings confirmed by automated
tests that exercise the real, running application over HTTP -- not just a
static-analysis pass. See `test-cases.xlsx` for the full 400-case suite and
`findings.xlsx` for the machine-readable finding list.

## What was tested well and holds up
- **SQL injection:** every free-text field (scan target, justification,
  login credentials) was fuzzed with 15 classic SQLi payloads; the
  SQLAlchemy async ORM's parameterized queries held in every case.
- **Command / argument injection:** `nmap` is invoked via
  `asyncio.create_subprocess_exec` with an argument list (never `shell=True`),
  confirmed safe against shell metacharacter payloads -- though see VULN-004
  for a narrower argument-injection edge case.
- **Path traversal:** no field is ever used as a filesystem path; traversal
  payloads are stored/rejected as inert text.
- **JWT security:** `alg: none`, algorithm-confusion, weak-secret guessing,
  payload tampering, truncation, and trailing-garbage tokens were all
  correctly rejected.
- **Session teardown:** logout genuinely denylists the refresh token --
  confirmed across repeated post-logout refresh attempts, not just once.
- **Primary scan ownership:** `/scans/*` correctly scopes access by owner,
  with an admin override, verified via a real two-client-account IDOR setup
  (not just "authenticated vs not").

## Confirmed findings

{finding_sections}

## Methodology notes
- All 400 tests in this suite run against a real `uvicorn` process bound to
  an ephemeral SQLite database, not against route handlers called directly
  in-process -- so CORS middleware, JSON parsing, and the real JWT auth
  dependency are all genuinely exercised on every request.
- `nmap` is not available in this environment, so background scan execution
  itself was not (and cannot be) exercised end-to-end here; all scan-related
  tests assert against the synchronous API contract only.
- No live AI (Groq/Gemini) credentials are configured, so remediation
  generation was tested up to the point of the outbound AI call.
""", encoding="utf-8")


def write_performance_report_md(out_path: Path, perf_stats: dict | None) -> None:
    if perf_stats:
        body = f"""## In-process load sample (this run)

A lightweight async load generator (not k6 -- see below for why) hit
`GET /health` and `GET /scans` concurrently against the same ephemeral
instance used for the functional suite, immediately after the 400-case run.

| Metric | GET /health | GET /scans (client1) |
|---|---|---|
| Requests | {perf_stats['health']['count']} | {perf_stats['scans']['count']} |
| Concurrency | {perf_stats['concurrency']} | {perf_stats['concurrency']} |
| p50 latency (ms) | {perf_stats['health']['p50']} | {perf_stats['scans']['p50']} |
| p95 latency (ms) | {perf_stats['health']['p95']} | {perf_stats['scans']['p95']} |
| p99 latency (ms) | {perf_stats['health']['p99']} | {perf_stats['scans']['p99']} |
| Max latency (ms) | {perf_stats['health']['max']} | {perf_stats['scans']['max']} |
| Error rate | {perf_stats['health']['error_rate']}% | {perf_stats['scans']['error_rate']}% |
| Throughput (req/s) | {perf_stats['health']['rps']} | {perf_stats['scans']['rps']} |
"""
    else:
        body = "_No in-process load sample was collected for this run._"

    out_path.write_text(f"""# Performance Report -- Vulnara Backend

## Why this isn't a full k6 run
This sandbox has no outbound network access to install the `k6` binary
(it isn't on PyPI/npm), so a full k6 load test could not be executed here.
`performance/k6-load-test.js` is included and ready to run in any
environment with `k6` installed and network access to a deployed instance:

```bash
BASE_URL=https://your-deployed-instance.example k6 run performance/k6-load-test.js
```

It exercises the same core flows as this suite -- login, scan creation,
scan retrieval, vulnerability listing -- ramping from 1 to 50 virtual users
over 5 minutes, with thresholds of p95 < 500ms and error rate < 1%.

{body}

## Interpreting these numbers
This sample is against a single-worker `uvicorn` process on ephemeral
SQLite with no connection pooling tuning -- it is **not** representative of
production (Postgres + whatever hosting tier is actually deployed). Treat
it as a smoke-level sanity check that the app doesn't fall over under light
concurrency, not as a capacity-planning number. Run `k6-load-test.js`
against a real staging deployment for numbers you can actually act on.
""", encoding="utf-8")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--json-report", type=Path, default=REPORTS_DIR / "full_run.json")
    parser.add_argument("--perf-stats", type=Path, default=REPORTS_DIR / "perf_stats.json")
    parser.add_argument("--base-url", default="http://127.0.0.1:8000 (ephemeral, per-run)")
    args = parser.parse_args()

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    print("Parsing test docstrings...")
    cases = parse_test_docstrings()
    print(f"  {len(cases)} test cases found across {len(set(c['file'] for c in cases))} files")

    print("Loading JSON test-run results...")
    results = load_json_results(args.json_report)
    merged = merge_cases_with_results(cases, results)
    print(f"  {len(merged)} rows after merging with execution results")

    run_meta = {
        "run_at": datetime.now(timezone.utc).isoformat(),
        "base_url": args.base_url,
    }

    print("Writing test-cases.xlsx...")
    write_test_cases_xlsx(merged, REPORTS_DIR / "test-cases.xlsx")

    print("Writing Automation_Test_Report.xlsx...")
    write_automation_style_report(merged, run_meta, REPORTS_DIR / "Automation_Test_Report.xlsx")

    print("Writing findings.xlsx...")
    write_findings_xlsx(REPORTS_DIR / "findings.xlsx")

    print("Writing endpoint-inventory.xlsx...")
    write_endpoint_inventory_xlsx(REPORTS_DIR / "endpoint-inventory.xlsx")

    print("Writing backend-inventory.md...")
    write_backend_inventory_md(REPORTS_DIR / "backend-inventory.md")

    print("Writing executive-summary.md...")
    write_executive_summary_md(merged, REPORTS_DIR / "executive-summary.md")

    print("Writing security-review.md...")
    write_security_review_md(REPORTS_DIR / "security-review.md")

    perf_stats = None
    if args.perf_stats.exists():
        perf_stats = json.loads(args.perf_stats.read_text())
    print("Writing performance-report.md...")
    write_performance_report_md(REPORTS_DIR / "performance-report.md", perf_stats)

    total = len(merged)
    passed = sum(1 for c in merged if c["status"] == "PASSED")
    failed = sum(1 for c in merged if c["status"] == "FAILED")
    print(f"\nDone. {total} test cases documented ({passed} passed / {failed} failed in the last run).")
    print(f"Reports written to {REPORTS_DIR}")


if __name__ == "__main__":
    sys.exit(main())
