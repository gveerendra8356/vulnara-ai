"""
reports/generate_reports.py

Runs as its own CI step AFTER pytest finishes (never inside a pytest hook --
see conftest.py's docstring). Reads reports/execution-results.json (written
by pytest_sessionfinish) and produces:

    reports/Automation_Test_Report.xlsx   (Executed Tests / Passed / Failed /
                                            Skipped / Execution Metrics /
                                            Defect Summary sheets)
    reports/dashboard.html                (pass-rate bar + per-module chart)
    reports/execution-report.html         (sortable-by-eye full test table)
    reports/summary.md                    (CI job summary / PR comment body)

Sheet layout, column widths, and the module -> severity mapping intentionally
mirror the reference Automation_Test_Report.xlsx supplied for this project so
the two report sets are visually and structurally consistent.
"""

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REPORTS_DIR = os.environ.get("REPORTS_DIR", os.path.dirname(os.path.abspath(__file__)))
RESULTS_PATH = os.path.join(REPORTS_DIR, "execution-results.json")
APP_NAME = "Vulnara"
GATE_THRESHOLD = float(os.environ.get("GATE_THRESHOLD", "90"))

MODULE_DISPLAY_NAMES = {
    "authentication": "Authentication",
    "authorization": "Authorization",
    "navigation": "Navigation",
    "ui_validation": "UI Validation",
    "forms": "Forms",
    "crud_operations": "CRUD Operations",
    "input_validation": "Input Validation",
    "error_handling": "Error Handling",
    "session_management": "Session Management",
    "search_filter": "Search & Filter",
    "accessibility": "Accessibility",
    "responsive": "Responsive",
}

MODULE_SEVERITY = {
    "Authentication": "HIGH",
    "Authorization": "HIGH",
    "CRUD Operations": "MEDIUM",
    "Forms": "MEDIUM",
    "Input Validation": "MEDIUM",
    "Error Handling": "MEDIUM",
    "Session Management": "MEDIUM",
    "Search & Filter": "LOW",
    "Accessibility": "LOW",
    "Navigation": "LOW",
    "Responsive": "LOW",
    "UI Validation": "LOW",
}

HEADER_FILL = PatternFill(start_color="00F0F0F0", end_color="00F0F0F0", fill_type="solid")
HEADER_FONT = Font(bold=True)


def module_display(raw: str) -> str:
    return MODULE_DISPLAY_NAMES.get(raw, raw.replace("_", " ").title())


def load_results():
    if not os.path.exists(RESULTS_PATH):
        print(f"No execution-results.json found at {RESULTS_PATH}", file=sys.stderr)
        sys.exit(1)
    with open(RESULTS_PATH) as f:
        return json.load(f)


def write_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def build_workbook(payload) -> Workbook:
    results = payload["results"]
    wb = Workbook()

    # ---------------------------------------------------------- Executed Tests
    ws = wb.active
    ws.title = "Executed Tests"
    write_header(ws, ["#", "Test ID", "Module", "Markers", "Status", "Duration (s)"])
    for i, r in enumerate(results, start=1):
        ws.append([i, r["test_id"], module_display(r["module"]), ",".join(r["markers"]),
                   r["status"], r["duration"]])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14

    # --------------------------------------------------------------- Passed
    ws_p = wb.create_sheet("Passed")
    write_header(ws_p, ["#", "Test ID", "Module", "Duration (s)"])
    passed = [r for r in results if r["status"] == "PASSED"]
    for i, r in enumerate(passed, start=1):
        ws_p.append([i, r["test_id"], module_display(r["module"]), r["duration"]])
    for col, w in zip("ABCD", (10, 90, 20, 14)):
        ws_p.column_dimensions[col].width = w

    # --------------------------------------------------------------- Failed
    ws_f = wb.create_sheet("Failed")
    write_header(ws_f, ["#", "Test ID", "Module", "Duration (s)"])
    failed = [r for r in results if r["status"] in ("FAILED", "ERROR")]
    for i, r in enumerate(failed, start=1):
        ws_f.append([i, r["test_id"], module_display(r["module"]), r["duration"]])
    for col, w in zip("ABCD", (10, 90, 20, 14)):
        ws_f.column_dimensions[col].width = w

    # -------------------------------------------------------------- Skipped
    ws_s = wb.create_sheet("Skipped")
    write_header(ws_s, ["#", "Test ID", "Module", "Duration (s)"])
    skipped = [r for r in results if r["status"] == "SKIPPED"]
    for i, r in enumerate(skipped, start=1):
        ws_s.append([i, r["test_id"], module_display(r["module"]), r["duration"]])
    for col, w in zip("ABCD", (10, 90, 20, 14)):
        ws_s.column_dimensions[col].width = w

    # ------------------------------------------------------ Execution Metrics
    ws_m = wb.create_sheet("Execution Metrics")
    write_header(ws_m, ["Metric", "Value"])
    total = payload["total"]
    pass_rate = payload["pass_rate"]
    gate_passed = "YES" if pass_rate >= GATE_THRESHOLD else "NO"
    for row in [
        ("Run At", payload["run_at"]),
        ("Base URL", payload["base_url"]),
        ("Total Tests", total),
        ("Passed", payload["passed"]),
        ("Failed", payload["failed"]),
        ("Skipped", payload["skipped"]),
        ("Errors", payload["errors"]),
        ("Pass Rate (%)", pass_rate),
        ("Gate Threshold (%)", GATE_THRESHOLD),
        ("Gate Passed", gate_passed),
        ("Total Duration (s)", payload["duration_seconds"]),
    ]:
        ws_m.append(row)
    ws_m.column_dimensions["A"].width = 20
    ws_m.column_dimensions["B"].width = 55

    # ------------------------------------------------------- Defect Summary
    ws_d = wb.create_sheet("Defect Summary")
    write_header(ws_d, ["#", "Defect / Test ID", "Module", "Severity"])
    for i, r in enumerate(failed, start=1):
        mod = module_display(r["module"])
        ws_d.append([i, r["test_id"], mod, MODULE_SEVERITY.get(mod, "MEDIUM")])
    for col, w in zip("ABCD", (10, 90, 20, 14)):
        ws_d.column_dimensions[col].width = w

    return wb


def build_summary_md(payload) -> str:
    results = payload["results"]
    by_module = {}
    for r in results:
        mod = r["module"]
        by_module.setdefault(mod, {"PASSED": 0, "FAILED": 0, "SKIPPED": 0, "ERROR": 0})
        by_module[mod][r["status"]] = by_module[mod].get(r["status"], 0) + 1

    gate = "PASSED" if payload["pass_rate"] >= GATE_THRESHOLD else "FAILED"
    lines = [
        f"# {APP_NAME} - Selenium Web Test Summary",
        "",
        f"- **Total executed:** {payload['total']}",
        f"- **Passed:** {payload['passed']}",
        f"- **Failed:** {payload['failed']}",
        f"- **Skipped:** {payload['skipped']}",
        f"- **Errors:** {payload['errors']}",
        f"- **Pass rate:** {payload['pass_rate']}% (threshold: {GATE_THRESHOLD}%)",
        f"- **Gate:** {gate}",
        f"- **Total duration:** {payload['duration_seconds']}s",
        f"- **Base URL:** {payload['base_url']}",
        "",
        "## By module",
        "",
        "| Module | Passed | Failed | Skipped |",
        "|---|---|---|---|",
    ]
    for mod in sorted(by_module):
        c = by_module[mod]
        lines.append(f"| `tests/test_{mod}.py` | {c.get('PASSED', 0)} | "
                      f"{c.get('FAILED', 0) + c.get('ERROR', 0)} | {c.get('SKIPPED', 0)} |")
    lines.append("")
    return "\n".join(lines)


def build_dashboard_html(payload) -> str:
    results = payload["results"]
    by_module = {}
    for r in results:
        mod = r["module"]
        by_module.setdefault(mod, {"PASSED": 0, "FAILED": 0, "SKIPPED": 0, "ERROR": 0})
        by_module[mod][r["status"]] = by_module[mod].get(r["status"], 0) + 1

    pass_rate = payload["pass_rate"]
    gate_class = "gate-pass" if pass_rate >= GATE_THRESHOLD else "gate-fail"

    bars = []
    for mod in sorted(by_module):
        c = by_module[mod]
        total = sum(c.values()) or 1
        p_pct = 100 * c.get("PASSED", 0) / total
        f_pct = 100 * (c.get("FAILED", 0) + c.get("ERROR", 0)) / total
        s_pct = 100 * c.get("SKIPPED", 0) / total
        bars.append(f"""<div style="margin-bottom:0.75rem">
  <div style="font-size:0.85rem;margin-bottom:0.2rem">test_{mod} ({total})</div>
  <div style="display:flex;height:14px;border-radius:7px;overflow:hidden;background:#eee">
    <div style="width:{p_pct}%;background:#1e8e4a"></div>
    <div style="width:{f_pct}%;background:#d64545"></div>
    <div style="width:{s_pct}%;background:#b58900"></div>
  </div>
</div>""")

    return f"""<!doctype html>
<html><head><meta charset="utf-8"><title>{APP_NAME} - Test Dashboard</title>
<style>
body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 2rem; }}
h1 {{ margin-bottom: 0.25rem; }}
.big {{ font-size: 3rem; font-weight: 700; }}
.gate-pass {{ color: #1e8e4a; }}
.gate-fail {{ color: #d64545; }}
</style></head>
<body>
<h1>{APP_NAME} - Test Dashboard</h1>
<p class="big {gate_class}">
  {pass_rate}%
</p>
<p>Gate: {GATE_THRESHOLD}% required to pass CI &middot;
   {payload['passed']} passed / {payload['failed']} failed / {payload['skipped']} skipped &middot;
   {payload['duration_seconds']}s total</p>
<h2>By module</h2>
{"".join(bars)}
</body></html>"""


def build_execution_report_html(payload) -> str:
    results = payload["results"]
    rows = []
    for r in results:
        color = {"PASSED": "#1e8e4a", "FAILED": "#d64545", "SKIPPED": "#b58900", "ERROR": "#d64545"}.get(
            r["status"], "#333")
        rows.append(
            f"<tr><td>tests/test_{r['module']}.py::{r['full_name']}</td>"
            f"<td style=\"color:{color}\">{r['status']}</td><td>{r['duration']}</td></tr>"
        )

    return f"""<!doctype html>
<html><head><meta charset="utf-8">
<title>{APP_NAME} - Selenium Execution Report</title>
<style>
body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; margin: 2rem; color: #1a1a1a; }}
h1 {{ margin-bottom: 0.25rem; }}
.meta {{ color: #666; margin-bottom: 1.5rem; }}
.cards {{ display: flex; gap: 1rem; margin-bottom: 2rem; flex-wrap: wrap; }}
.card {{ padding: 1rem 1.5rem; border-radius: 10px; background: #f4f4f4; min-width: 120px; }}
.card b {{ display:block; font-size: 1.6rem; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ text-align: left; padding: 0.5rem 0.75rem; border-bottom: 1px solid #eee; font-size: 0.9rem; }}
th {{ background: #fafafa; position: sticky; top: 0; }}
</style></head>
<body>
<h1>{APP_NAME} - Selenium Execution Report</h1>
<p class="meta">Target: {payload['base_url']} &middot; Generated: {payload['run_at']}</p>
<div class="cards">
  <div class="card">Total<b>{payload['total']}</b></div>
  <div class="card" style="background:#eaf7ee">Passed<b style="color:#1e8e4a">{payload['passed']}</b></div>
  <div class="card" style="background:#fdeceb">Failed<b style="color:#d64545">{payload['failed']}</b></div>
  <div class="card" style="background:#fdf3d9">Skipped<b style="color:#b58900">{payload['skipped']}</b></div>
  <div class="card">Errors<b>{payload['errors']}</b></div>
  <div class="card">Pass rate<b>{payload['pass_rate']}%</b></div>
  <div class="card">Duration<b>{payload['duration_seconds']}s</b></div>
</div>
<table>
<thead><tr><th>Test</th><th>Status</th><th>Duration</th></tr></thead>
<tbody>
{"".join(rows)}
</tbody>
</table>
</body></html>"""


def main():
    payload = load_results()

    wb = build_workbook(payload)
    xlsx_path = os.path.join(REPORTS_DIR, "Automation_Test_Report.xlsx")
    wb.save(xlsx_path)
    print(f"Wrote {xlsx_path}")

    with open(os.path.join(REPORTS_DIR, "summary.md"), "w") as f:
        f.write(build_summary_md(payload))
    print("Wrote summary.md")

    with open(os.path.join(REPORTS_DIR, "dashboard.html"), "w") as f:
        f.write(build_dashboard_html(payload))
    print("Wrote dashboard.html")

    with open(os.path.join(REPORTS_DIR, "execution-report.html"), "w") as f:
        f.write(build_execution_report_html(payload))
    print("Wrote execution-report.html")

    pass_rate = payload["pass_rate"]
    if pass_rate < GATE_THRESHOLD:
        print(f"Quality gate FAILED: {pass_rate}% < {GATE_THRESHOLD}% threshold", file=sys.stderr)
        sys.exit(1)
    print(f"Quality gate PASSED: {pass_rate}% >= {GATE_THRESHOLD}%")


if __name__ == "__main__":
    main()
